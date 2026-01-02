import streamlit as st
import pandas as pd
from google.oauth2 import service_account
import gspread
from gspread_dataframe import set_with_dataframe
from datetime import datetime

# --- KONFIGURASI HALAMAN ---
st.set_page_config(
    page_title="Bali Trip Budget Planner",
    page_icon="🌴",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- CSS STYLING PROFESIONAL ---
st.markdown("""
<style>
    /* Main Title */
    .main-title {
        text-align: center;
        font-family: 'Helvetica Neue', sans-serif;
        color: #1E88E5;
        font-weight: 700;
        margin-bottom: 5px;
    }
    .subtitle {
        text-align: center;
        color: #757575;
        font-size: 16px;
        margin-bottom: 25px;
    }
    
    /* Metrics Cards */
    div[data-testid="metric-container"] {
        background-color: #FFFFFF;
        border: 1px solid #E0E0E0;
        padding: 20px;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        text-align: center;
        transition: transform 0.2s;
    }
    div[data-testid="metric-container"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.1);
    }
    
    /* Labels inside metrics */
    div[data-testid="metric-container"] label {
        font-size: 14px;
        color: #616161;
    }
    
    /* Values inside metrics */
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] {
        font-size: 24px;
        font-weight: bold;
        color: #212121;
    }

    /* Custom success/error boxes */
    .stAlert {
        border-radius: 8px;
    }
    
    /* Save Button container */
    .save-btn-container {
        text-align: right;
        margin-top: 10px;
        margin-bottom: 20px;
    }
</style>
""", unsafe_allow_html=True)

# --- KONFIGURASI GOOGLE SHEETS ---
SPREADSHEET_ID = "1TQAOaIcGsW9SiXySWXhpsABHkMsrPe1yf9x9a9FIZys"
WORKSHEET_NAME = "Sheet1"

@st.cache_resource
def init_gsheet_connection():
    """Initialize Google Sheets connection"""
    try:
        credentials_dict = st.secrets["gcp_service_account"]
        credentials = service_account.Credentials.from_service_account_info(
            credentials_dict,
            scopes=["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        )
        gc = gspread.authorize(credentials)
        return gc
    except Exception as e:
        st.error(f"❌ Gagal koneksi ke Google Sheets: {str(e)}")
        return None

@st.cache_data(ttl=60)
def load_data_from_sheet():
    """Load data from Google Sheet"""
    try:
        gc = init_gsheet_connection()
        if gc is None: return None
        
        spreadsheet = gc.open_by_key(SPREADSHEET_ID)
        worksheet = spreadsheet.worksheet(WORKSHEET_NAME)
        
        data = worksheet.get_all_values()
        
        # Define Columns (Update: Added Status)
        expected_cols = ['Nama Barang', 'Qty', 'Harga Input', 'Total Akhir', 'Tipe', 'Status']
        
        if len(data) <= 1:
            return pd.DataFrame(columns=expected_cols)
        
        # Load Data
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # Ensure 'Status' column exists (migration handling)
        if 'Status' not in df.columns:
            df['Status'] = "FALSE"
        
        # Convert Types
        df['Qty'] = pd.to_numeric(df['Qty'], errors='coerce').fillna(0).astype(int)
        df['Harga Input'] = pd.to_numeric(df['Harga Input'], errors='coerce').fillna(0).astype(int)
        df['Total Akhir'] = pd.to_numeric(df['Total Akhir'], errors='coerce').fillna(0).astype(int)
        
        # Convert Status to Boolean for Checkbox
        df['Status'] = df['Status'].apply(lambda x: True if str(x).upper() == 'TRUE' else False)
        
        # Reorder columns just in case
        df = df[expected_cols]
        
        return df
    
    except Exception as e:
        st.error(f"❌ Gagal load data: {str(e)}")
        return None

def append_to_sheet(nama_barang, qty, harga, total_akhir, tipe, status):
    """Append new row to Google Sheet"""
    try:
        gc = init_gsheet_connection()
        if gc is None: return False
        
        spreadsheet = gc.open_by_key(SPREADSHEET_ID)
        worksheet = spreadsheet.worksheet(WORKSHEET_NAME)
        
        # Convert boolean status to string for Sheets
        status_str = "TRUE" if status else "FALSE"
        new_row = [nama_barang, int(qty), int(harga), int(total_akhir), tipe, status_str]
        
        worksheet.append_row(new_row)
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"❌ Gagal menyimpan: {str(e)}")
        return False

def update_sheet_data(df_edited):
    """Overwrite worksheet with edited dataframe"""
    try:
        gc = init_gsheet_connection()
        if gc is None: return False
        
        spreadsheet = gc.open_by_key(SPREADSHEET_ID)
        worksheet = spreadsheet.worksheet(WORKSHEET_NAME)
        
        # Prepare DF for upload (Convert boolean back to string if needed, or let set_with_dataframe handle it)
        df_upload = df_edited.copy()
        df_upload['Status'] = df_upload['Status'].astype(str).str.upper()
        
        # Clear and Update
        worksheet.clear()
        set_with_dataframe(worksheet, df_upload)
        
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"❌ Gagal update data: {str(e)}")
        return False

# --- HEADER ---
st.markdown('<h1 class="main-title">🌴 Bali Trip Budget Planner</h1>', unsafe_allow_html=True)
st.markdown('<p class="subtitle">Kelola budget liburan Anda dengan mudah dan profesional</p>', unsafe_allow_html=True)

# --- SIDEBAR: INPUT ---
with st.sidebar:
    st.header("📝 Input Pengeluaran")
    st.info("Tambahkan rencana pengeluaran baru di sini.")
    
    with st.form("form_tambah_item", clear_on_submit=True):
        nama_barang = st.text_input("Nama Barang / Aktivitas", placeholder="Contoh: Tiket Pesawat")
        
        col_side1, col_side2 = st.columns(2)
        with col_side1:
            qty = st.number_input("Jumlah (Qty)", min_value=1, value=1)
        with col_side2:
            harga = st.number_input("Harga (Rp)", min_value=0, value=0, step=10000)
        
        tipe_harga = st.radio("Metode Hitung", ["Harga Satuan (x Qty)", "Harga Borongan (Total)"])
        is_purchased = st.checkbox("Sudah Dibayar/Dibeli?", value=False)
        
        submitted = st.form_submit_button("➕ Tambah Item", use_container_width=True)

    if submitted:
        if not nama_barang:
            st.warning("⚠️ Nama barang wajib diisi.")
        else:
            total_akhir = (harga * qty) if tipe_harga == "Harga Satuan (x Qty)" else harga
            tipe_str = "Satuan" if tipe_harga == "Harga Satuan (x Qty)" else "Borongan"
            
            if append_to_sheet(nama_barang, qty, harga, total_akhir, tipe_str, is_purchased):
                st.success(f"✅ {nama_barang} berhasil ditambahkan!")
                st.rerun()

# --- MAIN CONTENT ---

# 1. Load Data
df = load_data_from_sheet()

if df is not None:
    # 2. Calculate Metrics
    if not df.empty:
        total_estimasi = df['Total Akhir'].sum()
        
        # Filter based on Status (True = Terpakai)
        uang_terpakai = df[df['Status'] == True]['Total Akhir'].sum()
        uang_dibutuhkan = df[df['Status'] == False]['Total Akhir'].sum()
        
        # Calculate Percentage
        persen_terpakai = (uang_terpakai / total_estimasi * 100) if total_estimasi > 0 else 0
    else:
        total_estimasi = 0
        uang_terpakai = 0
        uang_dibutuhkan = 0
        persen_terpakai = 0

    # 3. Display Dashboard Metrics
    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("💰 Total Estimasi Budget", f"Rp {total_estimasi:,.0f}", help="Total seluruh rencana pengeluaran")
    with m2:
        st.metric("💸 Uang Terpakai (Paid)", f"Rp {uang_terpakai:,.0f}", f"{persen_terpakai:.1f}% dari total")
    with m3:
        st.metric("⏳ Sisa Uang Dibutuhkan", f"Rp {uang_dibutuhkan:,.0f}", help="Dana yang harus disiapkan untuk item yang belum dibeli", delta_color="inverse")

    st.markdown("---")

    # 4. Interactive Data Editor
    col_header, col_refresh = st.columns([4, 1])
    with col_header:
        st.subheader("📋 Rincian & Checklist")
    with col_refresh:
        if st.button("🔄 Refresh", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    if not df.empty:
        # Edit Data Config
        edited_df = st.data_editor(
            df,
            column_config={
                "Status": st.column_config.CheckboxColumn(
                    "Sudah Dibeli?",
                    help="Centang jika barang sudah dibeli/dibayar",
                    default=False,
                ),
                "Nama Barang": st.column_config.TextColumn("Nama Item", width="large"),
                "Qty": st.column_config.NumberColumn("Qty", width="small"),
                "Harga Input": st.column_config.NumberColumn("Harga Input", format="Rp %d"),
                "Total Akhir": st.column_config.NumberColumn("Total", format="Rp %d", disabled=True),
                "Tipe": st.column_config.TextColumn("Tipe", width="small", disabled=True),
            },
            hide_index=True,
            use_container_width=True,
            num_rows="dynamic", # Allow adding/deleting rows directly
            key="data_editor"
        )
        
        # Save Button Logic
        # We check if the edited dataframe is different from original to show save button or process logic
        # But Streamlit data_editor returns the full new dataframe state.
        
        col_left, col_right = st.columns([3, 1])
        with col_right:
            if st.button("💾 Simpan Perubahan Checklist", type="primary", use_container_width=True):
                # Recalculate totals for edited rows (in case user changed price directly in table)
                # Note: Logic below assumes user might edit qty/price in table too
                # If you want strict logic, recalculate 'Total Akhir' here before saving
                
                # Simple recalculation logic
                # Loop isn't ideal for large data but safe for list
                for index, row in edited_df.iterrows():
                     if row['Tipe'] == 'Satuan':
                         edited_df.at[index, 'Total Akhir'] = row['Harga Input'] * row['Qty']
                     else:
                         edited_df.at[index, 'Total Akhir'] = row['Harga Input']

                if update_sheet_data(edited_df):
                    st.success("✅ Data berhasil diupdate ke Google Sheets!")
                    st.rerun()

        # 5. Export Section
        st.markdown("### 📥 Export Laporan")
        col_ex1, col_ex2 = st.columns(2)
        
        # CSV Export
        csv = edited_df.to_csv(index=False).encode('utf-8-sig')
        col_ex1.download_button(
            "📄 Download CSV",
            csv,
            f"Budget_Bali_{datetime.now().strftime('%Y%m%d')}.csv",
            "text/csv",
            use_container_width=True
        )

        # Excel Export
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Budget Plan"
            
            # Headers with Style
            headers = edited_df.columns.tolist()
            ws.append(headers)
            
            header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Data
            for r in dataframe_to_rows(edited_df, index=False, header=False):
                ws.append(r)
                
            wb.save(output)
            col_ex2.download_button(
                "📊 Download Excel",
                output.getvalue(),
                f"Budget_Bali_{datetime.now().strftime('%Y%m%d')}.xlsx",
                use_container_width=True
            )
        except:
             # Fallback if openpyxl issues or simple approach needed
             pass

    else:
        st.info("Belum ada data. Silakan input data melalui sidebar.")
